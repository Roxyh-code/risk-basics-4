"""
Create a synthetic version of the SPY Rolling Option Portfolio Excel file.
Replaces all numerical price/value data with synthetic but realistic-looking data
while preserving all structure (sheet names, column headers, option names, dates, etc.).
"""

import openpyxl
import numpy as np
import random
from copy import copy
from datetime import datetime, date

# ── reproducibility ──────────────────────────────────────────────────────────
np.random.seed(42)
random.seed(42)

SRC = r"C:\Users\Yang\Desktop\data\risk\week4\Assignment 4 - SPY Rolling Option Portfolio - Student.xlsx"
DST = r"C:\Users\Yang\Desktop\data\risk\week4\Assignment 4 - SPY Rolling Option Portfolio - Student_SYNTHETIC.xlsx"

print("Loading workbook (data_only=True) …")
wb_src = openpyxl.load_workbook(SRC, data_only=True)
print(f"  Sheets: {wb_src.sheetnames}")

# ── helpers ──────────────────────────────────────────────────────────────────

def is_num(v):
    return isinstance(v, (int, float)) and v is not None

def scale_price(v, factor, noise_std=0.02):
    """Scale a non-negative price by factor and add multiplicative noise."""
    if v is None or not is_num(v):
        return v
    new_v = v * factor * np.random.lognormal(0, noise_std)
    return max(0.0, new_v)

def read_sheet_data(ws):
    """Return list-of-lists (rows × cols) for a worksheet."""
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(list(row))
    return data

# ── read all sheets ──────────────────────────────────────────────────────────
print("Reading all sheet data …")
sheets = {}
for name in wb_src.sheetnames:
    sheets[name] = read_sheet_data(wb_src[name])
    print(f"  {name}: {len(sheets[name])} rows × {len(sheets[name][0]) if sheets[name] else 0} cols")

# ═══════════════════════════════════════════════════════════════════════════════
# 1. CONFIG
# ═══════════════════════════════════════════════════════════════════════════════
print("\n[config] Modifying parameters …")
cfg = sheets["config"]
# cfg rows: row[0]=header?, then key-value pairs
# We need to find AUM and Notional rows by key name and change values
for row in cfg:
    if row[0] == "Portfolio AUM":
        row[1] = 95_000_000
        print(f"  Portfolio AUM -> 95,000,000")
    elif row[0] == "Notional of options bought at roll (per leg)":
        row[1] = 240_000_000
        print(f"  Notional -> 240,000,000")
    # Funding cost stays 0, OTM stays 0.2, AS_OF_DATE stays unchanged

# ═══════════════════════════════════════════════════════════════════════════════
# 2. HOLIDAYS and EXPIRY – keep as-is
# ═══════════════════════════════════════════════════════════════════════════════
print("[holidays] Kept as-is")
print("[expiry]   Kept as-is")

# ═══════════════════════════════════════════════════════════════════════════════
# 3. DATA SPOT – synthetic SPY and SOFR random walk
# ═══════════════════════════════════════════════════════════════════════════════
print("\n[Data Spot] Generating synthetic spot prices …")
spot_data = sheets["Data Spot"]
# Row 0: headers (None, 'BTSISOFR Index', None, 'SPY US Equity')
# Row 1: sub-headers (None, 'PX_Last', None, 'PX_Last')
# Row 2+: date, sofr, None, spy

# Find data rows (skip first 2 header rows)
n_spot_rows = len(spot_data) - 2  # number of actual data rows

# Generate synthetic SPY random walk
spy_start = 480.0
spy_returns = np.random.normal(0.0005, 0.008, n_spot_rows)
spy_prices = spy_start * np.cumprod(1 + spy_returns)

# Generate synthetic SOFR (small drift + noise around 111.46 level)
sofr_start = 111.46
sofr_noise = np.random.normal(0, 0.003, n_spot_rows)  # ~0.3% noise
# small downward drift to make it realistic
sofr_drift = np.linspace(0, -0.5, n_spot_rows)
sofr_prices = sofr_start + sofr_drift + np.cumsum(sofr_noise)
sofr_prices = np.maximum(sofr_prices, 100.0)  # floor

for i, row_idx in enumerate(range(2, len(spot_data))):
    row = spot_data[row_idx]
    # col 1 = SOFR, col 3 = SPY
    if row[1] is not None:
        row[1] = round(float(sofr_prices[i]), 4)
    if row[3] is not None:
        row[3] = round(float(spy_prices[i]), 4)

print(f"  SPY range: {spy_prices.min():.2f} – {spy_prices.max():.2f}")
print(f"  SOFR range: {sofr_prices.min():.4f} – {sofr_prices.max():.4f}")

# Build lookup: date -> (sofr, spy)  for use in Leg sheets
# Dates are in col 0
spot_by_date = {}
for row_idx in range(2, len(spot_data)):
    row = spot_data[row_idx]
    d = row[0]  # could be datetime or date
    if d is not None:
        spot_by_date[d] = (row[1], row[3])  # (sofr, spy)

# ═══════════════════════════════════════════════════════════════════════════════
# 4. DATA OPTIONS BID / ASK – per-column multipliers
# ═══════════════════════════════════════════════════════════════════════════════
print("\n[Data Options Bid/Ask] Generating per-option multipliers …")
bid_data = sheets["Data Options Bid"]
ask_data = sheets["Data Options Ask"]

# Headers are in rows 0 and 1; data starts at row 2
# Col 0 = date; cols 1..N = option prices
n_option_cols = len(bid_data[0]) - 1  # number of option columns
print(f"  Number of option columns: {n_option_cols}")

# Generate one lognormal multiplier per option column (same for bid and ask)
option_multipliers = np.random.lognormal(0.0, 0.15, n_option_cols)
print(f"  Multiplier range: {option_multipliers.min():.3f} – {option_multipliers.max():.3f}")

def transform_option_sheet(data, multipliers, spread_extra=0.0):
    """Transform option price data in-place. Row 0,1 = headers. Col 0 = date."""
    for row in data[2:]:  # skip header rows
        for col_idx in range(1, len(row)):
            v = row[col_idx]
            if v is None or not is_num(v):
                continue
            m = multipliers[col_idx - 1]
            # multiplicative noise per cell
            noise = np.random.lognormal(0, 0.03)
            new_v = v * m * noise + spread_extra
            row[col_idx] = max(0.0, round(new_v, 4))

# Transform bid
print("  Transforming bid prices …")
transform_option_sheet(bid_data, option_multipliers, spread_extra=0.0)

# Transform ask with SAME multipliers but slightly wider spread
print("  Transforming ask prices …")
transform_option_sheet(ask_data, option_multipliers, spread_extra=0.0)

# Enforce ask >= bid cell by cell
print("  Enforcing ask >= bid …")
for r_idx in range(2, len(bid_data)):
    b_row = bid_data[r_idx]
    a_row = ask_data[r_idx]
    for c_idx in range(1, len(b_row)):
        bv = b_row[c_idx]
        av = a_row[c_idx]
        if bv is not None and av is not None and is_num(bv) and is_num(av):
            # Add a small random spread (0.5% – 2%) on top of bid for ask
            spread = bv * np.random.uniform(0.005, 0.02)
            a_row[c_idx] = round(max(av, bv + spread), 4)

# ═══════════════════════════════════════════════════════════════════════════════
# 5. LEG A / B / C
# ═══════════════════════════════════════════════════════════════════════════════
# Column indices (0-based):
# 0=DATE, 1=LEG, 2=OPTION_NAME, 3=QUANTITY, 4=PX_BID_CURR, 5=PX_MID_CURR,
# 6=PX_ASK_CURR, 7=PX_BID_PREV, 8=UNDERLYING_PX_LAST, 9=SOFR_TRI, 10=SOFR_DTD,
# 11=PREMIUM, 12=PROCEED, 13=OPTION_EXPIRY, 14=OPTION_STRIKE,
# 15=None, 16=ROLL?, 17=ROLL_OPTION, 18=ROLL_EXPIRY, 19=ROLL_STRIKE,
# 20=ROLL_QUANTITY, 21=None, 22=PNL_OPTION, 23=PNL_FUNDING, 24=PNL_TOTAL,
# 25=PNL_TOTAL_CUMSUM

def normalize_date(d):
    """Normalize date/datetime to date for lookup."""
    if isinstance(d, datetime):
        return d.date()
    if isinstance(d, date):
        return d
    return d

def transform_leg(leg_name, data):
    print(f"\n[{leg_name}] Transforming {len(data)-1} data rows …")
    # row 0 = header
    header = data[0]

    # Per-leg quantity factor
    qty_factor = np.random.uniform(0.85, 1.15)
    print(f"  Quantity factor: {qty_factor:.4f}")

    # Per-leg price scale factor (for option prices)
    price_factor = np.random.lognormal(0.0, 0.1)
    print(f"  Price factor: {price_factor:.4f}")

    pnl_totals = []

    for row in data[1:]:
        # ── date lookup for spot prices ──────────────────────────────────────
        d = row[0]
        d_key = normalize_date(d)

        # Get synthetic spot values (SOFR and SPY) for this date
        spot = spot_by_date.get(d_key) or spot_by_date.get(d)
        spy_val = None
        sofr_val = None
        if spot:
            sofr_val, spy_val = spot

        # ── QUANTITY (col 3) ─────────────────────────────────────────────────
        if row[3] is not None and is_num(row[3]):
            row[3] = round(row[3] * qty_factor)

        # ── ROLL_QUANTITY (col 20) ───────────────────────────────────────────
        if row[20] is not None and is_num(row[20]):
            row[20] = round(row[20] * qty_factor)

        # ── OPTION PRICES: bid, mid, ask, prev_bid ───────────────────────────
        bid = row[4]
        mid = row[5]
        ask = row[6]
        prev_bid = row[7]

        noise_b = np.random.lognormal(0, 0.025)
        noise_a_extra = np.random.uniform(0.005, 0.02)  # ask spread

        new_bid = None
        new_mid = None
        new_ask = None
        new_prev = None

        if bid is not None and is_num(bid):
            new_bid = max(0.0, bid * price_factor * noise_b)
        if ask is not None and is_num(ask):
            noise_a = np.random.lognormal(0, 0.025)
            new_ask = max(0.0, ask * price_factor * noise_a)
        if new_bid is not None and new_ask is not None:
            new_ask = max(new_ask, new_bid * (1 + noise_a_extra))
            new_mid = (new_bid + new_ask) / 2.0
        elif new_bid is not None:
            new_mid = new_bid
        elif new_ask is not None:
            new_mid = new_ask

        if prev_bid is not None and is_num(prev_bid):
            noise_pb = np.random.lognormal(0, 0.025)
            new_prev = max(0.0, prev_bid * price_factor * noise_pb)

        row[4] = round(new_bid, 4) if new_bid is not None else None
        row[5] = round(new_mid, 4) if new_mid is not None else None
        row[6] = round(new_ask, 4) if new_ask is not None else None
        row[7] = round(new_prev, 4) if new_prev is not None else None

        # ── UNDERLYING_PX_LAST (col 8) ───────────────────────────────────────
        if spy_val is not None:
            row[8] = round(spy_val, 4)
        elif row[8] is not None and is_num(row[8]):
            row[8] = round(row[8] * price_factor, 4)

        # ── SOFR_TRI (col 9) ────────────────────────────────────────────────
        if sofr_val is not None:
            row[9] = round(sofr_val, 4)
        elif row[9] is not None and is_num(row[9]):
            noise_s = np.random.normal(1.0, 0.003)
            row[9] = round(row[9] * noise_s, 4)

        # ── SOFR_DTD (col 10) ────────────────────────────────────────────────
        if row[10] is not None and is_num(row[10]):
            noise_sd = np.random.normal(1.0, 0.002)
            row[10] = round(row[10] * noise_sd, 8)

        # ── PREMIUM (col 11) and PROCEED (col 12) ───────────────────────────
        for ci in (11, 12):
            if row[ci] is not None and is_num(row[ci]):
                noise_p = np.random.lognormal(0, 0.05)
                row[ci] = round(row[ci] * price_factor * qty_factor * noise_p, 2)

        # ── PNL columns (22=OPTION, 23=FUNDING, 24=TOTAL) ───────────────────
        pnl_scale = np.random.uniform(0.8, 1.2)
        for ci in (22, 23, 24):
            if row[ci] is not None and is_num(row[ci]):
                noise_pnl = np.random.normal(1.0, 0.05)
                row[ci] = round(row[ci] * pnl_scale * noise_pnl, 2)

        pnl_total = row[24]
        pnl_totals.append(pnl_total if (pnl_total is not None and is_num(pnl_total)) else 0.0)

        # col 25 (CUMSUM) will be recomputed below

    # ── Recompute PNL_TOTAL_CUMSUM ───────────────────────────────────────────
    cumsum = 0.0
    for i, row in enumerate(data[1:]):
        pnl = row[24]
        if pnl is not None and is_num(pnl):
            cumsum += pnl
        row[25] = round(cumsum, 2)

    print(f"  Final cumsum PnL: {cumsum:,.2f}")

transform_leg("Leg A", sheets["Leg A"])
transform_leg("Leg B", sheets["Leg B"])
transform_leg("Leg C", sheets["Leg C"])

# ═══════════════════════════════════════════════════════════════════════════════
# 6. LEGS ALL – scale PnL values ±20%
# ═══════════════════════════════════════════════════════════════════════════════
print("\n[Legs All] Scaling PnL values …")
legs_all = sheets["Legs All"]
# Row 0 = headers; remaining rows = data
# We'll scale all numerical cells in data rows
for row in legs_all[1:]:
    for ci in range(len(row)):
        v = row[ci]
        if v is not None and is_num(v):
            scale = np.random.uniform(0.8, 1.2)
            row[ci] = round(v * scale, 2)

# ═══════════════════════════════════════════════════════════════════════════════
# 7. OPTIONS ALL – scale numerical values
# ═══════════════════════════════════════════════════════════════════════════════
print("[Options All] Scaling numerical values …")
options_all = sheets["Options All"]
for row in options_all[1:]:
    for ci in range(len(row)):
        v = row[ci]
        if v is not None and is_num(v):
            scale = np.random.lognormal(0.0, 0.10)
            row[ci] = round(v * scale, 4)

# ═══════════════════════════════════════════════════════════════════════════════
# 8. OUTPUT – scale return values
# ═══════════════════════════════════════════════════════════════════════════════
print("[Output] Scaling return values …")
output_data = sheets["Output"]
# Row 0 = headers; col 0 = DATE; remaining = return values
for row in output_data[1:]:
    for ci in range(1, len(row)):
        v = row[ci]
        if v is not None and is_num(v):
            scale = np.random.uniform(0.9, 1.1)
            row[ci] = round(v * scale, 8)

# ═══════════════════════════════════════════════════════════════════════════════
# 9. WRITE OUTPUT WORKBOOK
# ═══════════════════════════════════════════════════════════════════════════════
print("\nCreating output workbook …")
wb_dst = openpyxl.Workbook()
wb_dst.remove(wb_dst.active)  # remove default sheet

# Copy sheet order from source
for sheet_name in wb_src.sheetnames:
    ws_src = wb_src[sheet_name]
    ws_dst = wb_dst.create_sheet(title=sheet_name)

    print(f"  Writing sheet: {sheet_name} …", end=" ")
    data = sheets[sheet_name]

    for r_idx, row in enumerate(data, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws_dst.cell(row=r_idx, column=c_idx, value=value)

    # Copy basic column dimensions from source for readability
    for col_letter, col_dim in ws_src.column_dimensions.items():
        ws_dst.column_dimensions[col_letter].width = col_dim.width

    print(f"{len(data)} rows written")

print(f"\nSaving to:\n  {DST}")
wb_dst.save(DST)
print("Save complete.")

# ═══════════════════════════════════════════════════════════════════════════════
# SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 60)
print("SUMMARY OF SYNTHETIC FILE CREATION")
print("=" * 60)
print(f"Source : {SRC}")
print(f"Output : {DST}")
print()
print("Sheets processed:")
print("  config        – Portfolio AUM: 100M->95M, Notional: 250M->240M")
print("  holidays      – kept as-is (date reference data)")
print("  expiry        – kept as-is (date reference data)")
print("  Data Spot     – SPY random walk from ~480; SOFR ~111.46 with noise")
print("  Data Options Bid – per-option lognormal multiplier + per-cell noise")
print("  Data Options Ask – same multiplier as bid, enforced ask>=bid+spread")
print("  Leg A         – prices/quantities/PnL scaled; cumsum recalculated")
print("  Leg B         – prices/quantities/PnL scaled; cumsum recalculated")
print("  Leg C         – prices/quantities/PnL scaled; cumsum recalculated")
print("  Legs All      – PnL values scaled ±20%")
print("  Options All   – numerical values scaled with lognormal noise")
print("  Output        – return values scaled ±10%")
print()
print("Constraints enforced:")
print("  - PX_BID <= PX_MID <= PX_ASK")
print("  - All option prices >= 0")
print("  - PNL_TOTAL_CUMSUM recomputed from PNL_TOTAL")
print("  - None/null cells preserved as None")
print("  - Dates, option names, strike/expiry info unchanged")
print("  - ROLL? flags unchanged")
print("Done.")
